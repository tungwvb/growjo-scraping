# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
import os,csv,re
from openpyxl import Workbook
from openpyxl.styles import Font
from crawldata.functions import *

class CrawldataPipeline:
    def open_spider(self,spider):
        self.EXPORT="EXCEL"   # TO: EXCEL / CSV
        self.SHEET_MAX_ROW=300000
        if not os.path.exists('./Data'):
            os.mkdir('./Data',0o777)
        self.HEADER=False
        self.fieldnames=[]
        self.filename=None
        if self.EXPORT=='EXCEL':
            self.wb=Workbook()
            self.SHEET={}
            self.SHEET_COUNT={}
            self.SHEET_No={}
        else:
            self.CSV_FILE_No=0
            self.CSV_ROW_No=0
    def close_spider(self,spider):
        if self.EXPORT=='CSV':
            if self.CSV_ROW_No<self.SHEET_MAX_ROW and len(self.fieldnames)>0 and not self.filename is None:
                self.filename.close()
        elif self.EXPORT=='EXCEL':
            for sheet in self.SHEET:
                ws=self.wb[sheet]
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    if length>100:
                        length=100
                    ws.column_dimensions[column_cells[0].column_letter].width = length+1      
            self.wb.save('./Data/'+str(spider.name)+'.xlsx')
    def process_item(self, item, spider):
        if self.EXPORT=='CSV':
            for key in item.keys():
                if not key in self.fieldnames:
                    self.fieldnames.append(key)
            if self.HEADER==False:
                if self.CSV_FILE_No==0:
                    self.filename='./Data/Data_'+spider.name+'.csv'
                else:
                    self.filename='./Data/Data_'+spider.name+'_'+str(self.CSV_FILE_No)+'.csv'
                #filename = open(filename, mode='w',encoding="utf-8",newline='')
                self.filename = open(self.filename, mode='w',encoding="utf-8-sig",newline='')
                # Write Header
                writer = csv.DictWriter(self.filename, fieldnames=self.fieldnames)
                writer.writeheader()
                self.HEADER=True
            else:
                writer = csv.DictWriter(self.filename, fieldnames=self.fieldnames)
    
            data_row={}
            for i in range(0,len(self.fieldnames),1):
                if self.fieldnames[i] in item:
                    TXT=str(item[self.fieldnames[i]])
                    if TXT=='None':
                        TXT=''
                    data_row[self.fieldnames[i]]=TXT
                else:
                    data_row[self.fieldnames[i]]=''
            writer.writerow(data_row)
            self.CSV_ROW_No+=1
            if self.CSV_ROW_No>=self.SHEET_MAX_ROW:
                self.CSV_FILE_No+=1
                self.CSV_ROW_No=0
                self.filename.close()
                self.HEADER=False
        elif self.EXPORT=='EXCEL':
            if not 'SHEET' in item:
                item['SHEET']=spider.name
            if not item['SHEET'] in self.SHEET_No:
                self.SHEET_No[item['SHEET']]=0
            if self.SHEET_No[item['SHEET']]>0:
                item['SHEET']=item['SHEET']+'~'+str(self.SHEET_No[str(item['SHEET']).split('~')[0]])
            wskpi=self.wb.active                
            if len(self.SHEET)==0:
                wskpi.title=item['SHEET']  
                self.SHEET[item['SHEET']]=[]
                self.SHEET_COUNT[item['SHEET']]=0
                
            elif not item['SHEET'] in self.SHEET:
                wskpi=self.wb.create_sheet()                
                wskpi.title=item['SHEET']
                self.SHEET[item['SHEET']]=[]
                self.SHEET_COUNT[item['SHEET']]=0
                
            else:
                wskpi=self.wb[item['SHEET']]
            # Write HEADER
            for k in item.keys():
                if k!='SHEET' and not k in self.SHEET[item['SHEET']]:
                    self.SHEET[item['SHEET']].append(k)
                    header_font = Font(bold=True)
                    O=wskpi.cell(column=len(self.SHEET[item['SHEET']]), row=1, value=k)
                    O.font = header_font
            Data=[]
            for key in self.SHEET[item['SHEET']]:
                if key in item:
                    VALUE=item[key]
                else:
                    VALUE=''
                if str(VALUE).lower()=='none':
                    VALUE=''
                TYPE=self.get_DataType(VALUE)
                if TYPE=='INT':
                    Data.append(int(str(VALUE).replace(',','')))
                elif TYPE=='FLOAT':
                    Data.append(float(str(VALUE).replace(',','')))
                else:
                    Data.append(str(VALUE))
            wskpi.append(Data)
            
            self.SHEET_COUNT[item['SHEET']]+=1
            if int(self.SHEET_COUNT[item['SHEET']])>=self.SHEET_MAX_ROW:
                self.SHEET_No[str(item['SHEET']).split('~')[0]]+=1
        return item
    def get_DataType(self,strtxt):
        strtxt=str(strtxt).strip()
        if Get_Number(strtxt)==strtxt and len(Get_Number(strtxt))>0:
            if '.' in strtxt and str(strtxt).count('.')==1:
                return 'FLOAT'
            elif not '.' in str(strtxt) and len(strtxt)<9:
                return 'INT'
            else:
                return 'TEXT'
        else:
            return 'TEXT'