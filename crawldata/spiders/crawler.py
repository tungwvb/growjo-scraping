import scrapy,json,os,platform,validators,re,requests
from crawldata.functions import *
from datetime import datetime
from openpyxl import load_workbook

class CrawlerSpider(scrapy.Spider):
    name = 'growjo'
    DATE_CRAWL=datetime.now().strftime('%Y-%m-%d')
    if platform.system()=='Linux':
        URL='file:////' + os.getcwd()+'/scrapy.cfg'
    else:
        URL='file:///' + os.getcwd()+'/scrapy.cfg'
    LS=["annual revenue","revenue per employee","total funding","pricing","current valuation"]
    def Sorting(self,lst):
        lst2 = sorted(lst, key=len, reverse=True)
        return lst2
    def start_requests(self):
        wb = load_workbook(filename = 'Companies.xlsx',data_only=True)
        for i in range(len(wb.sheetnames)):
            SHEET=wb.sheetnames[i]
            data = wb[SHEET]
            for i in range(2,data.max_row+1,1):
                item={}
                item['KEYWORD']=''
                item['KEYWORDS']=[]
                for j in range(1,data.max_column+1,1):
                    if not data.cell(i,j).value in item:
                        item[data.cell(1,j).value]=str(data.cell(i,j).value).strip()
                for K in self.LS:
                    item[(str(K).replace('annual ','')).title()]=''
                item['List of 10 Competitors']=''
                item['KEYWORDS'].append(item['COMPANY'])
                item['KEYWORDS'].append(str(item['COMPANY']).replace("&"," and "))
                yield scrapy.Request(self.URL,callback=self.parse_keyword,meta={'item':item}, dont_filter=True)
    def parse_keyword(self, response):
        item=response.meta['item']
        COMPANY=(str(item['COMPANY']).split('(')[0]).strip()
        item['KEYWORDS'].append(COMPANY)
        if str(COMPANY).endswith('.'):
            COMPANYN=" ".join(str(COMPANY).split()[:-1])
            if str(COMPANY).endswith('and'):
                COMPANYN=" ".join(str(COMPANY).split()[:-1])
            if str(COMPANYN).endswith(",") or str(COMPANYN).endswith("&"):
                COMPANYN=str(COMPANYN[:-1]).strip()
            item['KEYWORDS'].append(COMPANYN)
        if str(COMPANY).endswith('Corp.'):
            COMPANYN=str(COMPANY).replace('Corp.','Corporation')
            item['KEYWORDS'].append(COMPANYN)
        if str(COMPANY).endswith('and'):
            COMPANYN=" ".join(str(COMPANY).split()[:-1])

        COMPANY=str(COMPANY).replace("&"," and ")
        item['KEYWORDS'].append(COMPANY)
        if str(COMPANY).endswith('.'):
            COMPANYN=" ".join(str(COMPANY).split()[:-1])
            if str(COMPANY).endswith('and'):
                COMPANYN=" ".join(str(COMPANY).split()[:-1])
            if str(COMPANYN).endswith(",") or str(COMPANYN).endswith("&"):
                COMPANYN=str(COMPANYN[:-1]).strip()
            item['KEYWORDS'].append(COMPANYN)
        if str(COMPANY).endswith('Corp.'):
            COMPANYN=str(COMPANY).replace('Corp.','Corporation')
            item['KEYWORDS'].append(COMPANYN)
        if str(COMPANY).endswith('and'):
            COMPANYN=" ".join(str(COMPANY).split()[:-1])

        COMPANY_NAME=mining_company_name((str(item['COMPANY']).split('(')[0]).strip())
        item['KEYWORDS'].append(COMPANY_NAME['PRODUCT'])
        item['KEYWORDS'].append(re.split(' Co.| Cos.',COMPANY_NAME['PRODUCT'])[0])
        if validators.domain(COMPANY_NAME['PRODUCT'])==True:
            item['KEYWORDS'].append((str(COMPANY_NAME['PRODUCT']).replace('www.','')).split(".")[0])
        item['KEYWORDS'] = self.Sorting(list(dict.fromkeys(item['KEYWORDS'])))
        KEYWORD=None
        TT=0
        while KEYWORD is None and TT<len(item['KEYWORDS']):
            KEYWORD=self.find_keyword(item['KEYWORDS'][TT])
            TT+=1
        if KEYWORD:
            item['KEYWORD']=KEYWORD
            data_post = '{"preference":"query"}\n{"query":{"bool":{"must":[{"bool":{"must":{"bool":{"should":[{"multi_match":{"query":"'+item['KEYWORDS'][0]+'","fields":["name"],"type":"best_fields","operator":"or","fuzziness":1}},{"multi_match":{"query":"'+item['KEYWORDS'][0]+'","fields":["name"],"type":"phrase_prefix","operator":"or"}}],"minimum_should_match":"1"}}}}]}},"size":20}\n'
            yield scrapy.Request('https://es.growjo.com/saasnew/doc/_msearch',method="POST",body=data_post,headers=self.headers_json,meta={'item':item,'TT':0}, callback=self.find_keyword,dont_filter=True)
        else:
            yield(item)
    def find_keyword(self,response):
        item=response.meta['item']
        TT=response.meta['TT']
        keyword=item['KEYWORDS'][TT]
        KEYWORD=None
        try:
            DATA=json.loads(response.text)
            if 'responses' in DATA:
                if len(DATA['responses'])>0:
                    Data=DATA['responses'][0]['hits']['hits']
                    for row in Data:
                        if (str(row['_source']['name']).upper()==str(keyword).strip().upper() or str(row['_source']['name']).upper()==str(keyword).strip().upper()+" COMPANY") and KEYWORD is None:
                            KEYWORD=row['_source']['name']
        except:
            pass
        if KEYWORD:
            item['KEYWORD']=KEYWORD
            url="https://growjo.com/company/"+str(KEYWORD).replace(" ","_")
            yield scrapy.Request(url,callback=self.parse_data,meta={'item':item})
        elif TT<len(item['KEYWORDS'])-1:
            TT+=1
            data_post = '{"preference":"query"}\n{"query":{"bool":{"must":[{"bool":{"must":{"bool":{"should":[{"multi_match":{"query":"'+item['KEYWORDS'][TT]+'","fields":["name"],"type":"best_fields","operator":"or","fuzziness":1}},{"multi_match":{"query":"'+item['KEYWORDS'][TT]+'","fields":["name"],"type":"phrase_prefix","operator":"or"}}],"minimum_should_match":"1"}}}}]}},"size":20}\n'
            yield scrapy.Request('https://es.growjo.com/saasnew/doc/_msearch',method="POST",body=data_post,headers=self.headers_json,meta={'item':item,'TT':TT}, callback=self.find_keyword,dont_filter=True)
        else:
            item['CRAWLED']=1
            yield(item)
    def parse_data(self,response):
        item=response.meta['item']
        item['CRAWLED']=1
        Data=response.xpath('//div[contains(@class,"description")]//ul[1]/li')
        for row in Data:
            TXT="".join(row.xpath('.//text()').getall())
            for K in self.LS:
                if K in TXT:
                    item[(str(K).replace('annual ','')).title()]=TXT
        Data=response.xpath('//div[contains(@class,"description")]//table/tbody/tr')
        Competitors=[]
        for row in Data:
            TXT=row.xpath('.//text()').getall()
            TEXT=[]
            for rs in TXT:
                rs=str(rs).strip()
                if rs!='':
                    TEXT.append(rs)
            Competitors.append(" ".join(TEXT))
        item['List of 10 Competitors']="\n".join(Competitors[:10])
        print('\n =========================================')
        yield(item)
